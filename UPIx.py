import os
import datetime
import imaplib
import email
import re
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import platform
import subprocess
from email.header import decode_header
from email.utils import parsedate_to_datetime
from openpyxl import load_workbook

# =========================
# CREDENTIALS / CONSTANTS
# =========================

EMAIL_USER = "" # Replace with your Gmail
EMAIL_PASS = ""  # Use App Password if 2FA is enabled

DEFAULT_EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Expenses", "Feb 25.xlsx")
SHEET_NAME = "Feb 25"

LOG_FILE = os.path.join(os.path.dirname(__file__), "log.txt")
LAST_PROCESSED_FILE = os.path.join(os.path.dirname(__file__), "last_processed_time.txt")

EXPENSE_CATEGORIES = {
    0: "Skip",
    1: "Food",
    2: "Travel",
    3: "Rent & Electricity",
    4: "Grooming Expense",
    5: "EMI Expense",
    6: "Indore Expense",
    7: "Subscription Based Expense",
    8: "Clothing Expense",
    9: "Business Related Expense",
    10: "Donation Expense",
    11: "Personal Expense",
    12: "Other Expenses"
}

# =========================
# DARK MODE DETECTION
# =========================

def detect_dark_mode():
    """
    Return True if the system is in Dark Mode (Windows or macOS).
    Otherwise, return False (assume light mode).
    """
    system = platform.system()
    if system == "Darwin":
        # macOS detection
        try:
            result = subprocess.run(
                ["defaults", "read", "-g", "AppleInterfaceStyle"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )
            return (result.returncode == 0 and "Dark" in result.stdout)
        except:
            pass
        return False
    elif system == "Windows":
        # Windows detection
        try:
            import winreg
            with winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize"
            ) as key:
                apps_use_light_theme, _ = winreg.QueryValueEx(key, "AppsUseLightTheme")
                return (apps_use_light_theme == 0)
        except:
            return False
    return False

# =========================
# GMAIL / EMAIL FUNCTIONS
# =========================

def connect_gmail():
    """Connect to Gmail via IMAP, return the mail object or None if error."""
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(EMAIL_USER, EMAIL_PASS)
        mail.select("inbox")
        return mail
    except Exception as e:
        print(f"Error connecting to Gmail: {e}")
        return None

def get_last_processed_time():
    if not os.path.exists(LAST_PROCESSED_FILE):
        return None
    try:
        with open(LAST_PROCESSED_FILE, "r", encoding="utf-8") as f:
            text = f.read().strip()
            return datetime.datetime.fromisoformat(text)
    except Exception:
        return None

def set_last_processed_time(dt):
    try:
        with open(LAST_PROCESSED_FILE, "w", encoding="utf-8") as f:
            f.write(dt.isoformat())
    except Exception as e:
        print(f"Warning: could not save last processed time: {e}")

# ========================
# EXCEL UPDATE FUNCTION
# ========================

def update_excel(transactions, category_choices, excel_file, sheet_name):
    from datetime import datetime

    total_upi_amount = round(sum(txn["amount"] for txn in transactions), 2)
    timestamp = datetime.now().strftime("%d-%b-%Y %H:%M:%S")
    log_entries = []

    # Grab last run total from the log
    last_run_total = 0.0
    try:
        with open(LOG_FILE, "r", encoding="utf-8") as log_file:
            lines = log_file.readlines()
            for line in reversed(lines):
                if "Total Expense Added Today" in line:
                    last_run_total = round(float(line.split(":")[1].strip().replace("Rs.", "").strip()), 2)
                    break
    except (FileNotFoundError, ValueError):
        pass

    # If today's total is the same as last time, skip
    if total_upi_amount == last_run_total and total_upi_amount != 0:
        msg = (f"Expenses for this set of transactions appear to be already logged. "
               f"(Mail total: Rs.{total_upi_amount:.2f})")
        messagebox.showwarning("Skipped", msg)
        with open(LOG_FILE, "a", encoding="utf-8") as log_file:
            log_file.write(f"\n[{timestamp}] ⚠️ SKIPPED: Expenses already recorded. No duplicate booking.\n")
            log_file.write("\n" + "=" * 50 + "\n\n")
        return

    try:
        wb = load_workbook(excel_file)
        ws_main = wb[sheet_name]
        ws_daily = wb["Daily 2025"]
    except FileNotFoundError:
        messagebox.showerror("Error", f"Excel file not found:\n{excel_file}")
        return
    except KeyError:
        messagebox.showerror("Error", f"Sheet '{sheet_name}' or 'Daily 2025' not found in workbook.")
        return

    ws_main["O1"].value = datetime.now().strftime("%d-%b-%Y %I:%M%p")

    total_amount_added = 0
    total_amount_skipped = 0
    category_sums = {}
    max_email_datetime = None

    for i, txn in enumerate(transactions):
        chosen_cat = category_choices[i] if category_choices[i] else "Skip"

        if chosen_cat == "Skip":
            total_amount_skipped += txn["amount"]
            log_entries.append(
                f"[{timestamp}] SKIPPED: Rs.{txn['amount']:.2f} for '{txn['party_name']}' (Skip chosen)"
            )
            continue

        msg_dt = txn["email_datetime"]
        if msg_dt and (max_email_datetime is None or msg_dt > max_email_datetime):
            max_email_datetime = msg_dt

        if chosen_cat == "Food":
            date_str = txn["date"]
            try:
                day = int(date_str.split("-")[0])
                date_obj = datetime.strptime(date_str, "%d-%m-%y")
                month_name = date_obj.strftime("%B")
            except:
                total_amount_skipped += txn["amount"]
                log_entries.append(
                    f"[{timestamp}] SKIPPED (Invalid date format '{date_str}'): Rs.{txn['amount']:.2f}"
                )
                continue

            # find the column in ws_daily for month_name
            month_col = None
            for col in range(2, ws_daily.max_column + 1):
                if ws_daily.cell(row=2, column=col).value == month_name:
                    month_col = col
                    break
            if not month_col:
                total_amount_skipped += txn["amount"]
                log_entries.append(
                    f"[{timestamp}] SKIPPED (No '{month_name}' column in 'Daily 2025'): Rs.{txn['amount']:.2f}"
                )
                continue

            # find the row for the day
            day_row = None
            for row_idx in range(3, ws_daily.max_row + 1):
                val = ws_daily.cell(row=row_idx, column=1).value
                if isinstance(val, (int, float)) and val == day:
                    day_row = row_idx
                    break
            if not day_row:
                total_amount_skipped += txn["amount"]
                log_entries.append(
                    f"[{timestamp}] SKIPPED (Day '{day}' not found in 'Daily 2025'): Rs.{txn['amount']:.2f}"
                )
                continue

            prev_food_exp = float(ws_daily.cell(row=day_row, column=month_col).value or 0)
            new_food_exp = prev_food_exp + txn["amount"]
            ws_daily.cell(row=day_row, column=month_col, value=new_food_exp)

            total_amount_added += txn["amount"]
            category_sums["Food"] = category_sums.get("Food", 0) + txn["amount"]
            log_entries.append(
                f"[{timestamp}] 'Daily 2025' (Food) -> Prev: Rs.{prev_food_exp:.2f}, +Rs.{txn['amount']:.2f}, New: Rs.{new_food_exp:.2f}"
            )
        else:
            # Non-Food
            category_row = None
            for row_idx in range(1, ws_main.max_row + 1):
                if ws_main.cell(row=row_idx, column=1).value == chosen_cat:
                    category_row = row_idx
                    break

            if category_row is None:
                total_amount_skipped += txn["amount"]
                log_entries.append(
                    f"[{timestamp}] SKIPPED (Category '{chosen_cat}' not found in main sheet): Rs.{txn['amount']:.2f}"
                )
                continue

            cell_value = ws_main.cell(row=category_row, column=3).value
            prev_balance = 0.0

            if isinstance(cell_value, str) and cell_value.startswith("="):
                updated_formula = f"{cell_value} + {txn['amount']}"
                ws_main.cell(row=category_row, column=3, value=updated_formula)
                try:
                    prev_str = cell_value[1:].strip()
                    prev_balance = float(eval(prev_str))
                except:
                    prev_balance = 0.0
                new_balance = prev_balance + txn["amount"]
            else:
                prev_balance = float(cell_value) if isinstance(cell_value, (int, float)) else 0.0
                new_balance = prev_balance + txn["amount"]
                ws_main.cell(row=category_row, column=3, value=new_balance)

            total_amount_added += txn["amount"]
            category_sums[chosen_cat] = category_sums.get(chosen_cat, 0) + txn["amount"]
            log_entries.append(
                f"[{timestamp}] '{chosen_cat}' -> Prev: Rs.{prev_balance:.2f}, +Rs.{txn['amount']:.2f}, New: Rs.{new_balance:.2f}"
            )

    wb.save(excel_file)
    wb.close()

    summary_msg = (
        f"Total UPI from Mail: Rs.{total_upi_amount:.2f}\n"
        f"Total Added: Rs.{total_amount_added:.2f}\n"
        f"Total Skipped: Rs.{total_amount_skipped:.2f}\n"
    )
    messagebox.showinfo("Summary", summary_msg)

    # Write logs with UTF-8 encoding
    with open(LOG_FILE, "a", encoding="utf-8") as lf:
        lf.write("\n" + "=" * 50 + "\n")
        for entry in log_entries:
            lf.write(entry + "\n")
        lf.write(f"\nTotal Expense Added Today: Rs. {total_amount_added:.2f}\n")
        lf.write(f"Total Amount Skipped: Rs.{total_amount_skipped:.2f}\n")
        lf.write("\n📂 **Category-wise Breakdown:**\n")
        for cat, amount in category_sums.items():
            lf.write(f"   - {cat}: Rs.{amount:.2f}\n")
        lf.write("\n" + "=" * 50 + "\n\n")

    if max_email_datetime:
        set_last_processed_time(max_email_datetime)

# ========================
# MAIN GUI
# ========================

class ExpenseGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        
        # ------------------------------
        # 1) Set the WINDOW title
        # ------------------------------
        self.title("💰 UPIx - A Simple UPI Expense Tracker")
        self.geometry("1000x700")

        # Decide theme by OS
        system = platform.system()
        style = ttk.Style(self)
        if system == "Windows":
            style.theme_use("default")
        elif system == "Darwin":
            style.theme_use("aqua")
        else:
            # Fallback for Linux/other systems
            style.theme_use("clam")

        # Decide text color for dark/light mode
        is_dark = detect_dark_mode()
        label_fg = "white" if is_dark else "black"

        style.configure("LastRun.TLabel", foreground=label_fg)

        # ------------------------------
        # 2) Add a BIG LABEL on top (centered)
        # ------------------------------
        title_label = ttk.Label(
            self, 
            text="💰 UPIx - A Simple UPI Expense Tracker", 
            font=("Helvetica", 16, "bold"),
            foreground=label_fg
        )
        title_label.pack(pady=10)

        # -- The only change: remove text="Excel File" from LabelFrame below --
        file_frame = ttk.LabelFrame(self)
        file_frame.pack(fill="x", padx=10, pady=10)

        ttk.Label(file_frame, text="Excel Path:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.excel_path_var = tk.StringVar(value=DEFAULT_EXCEL_FILE)
        ttk.Entry(file_frame, textvariable=self.excel_path_var, width=55).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_excel).grid(row=0, column=2, padx=5, pady=5)

        ttk.Label(file_frame, text="Excel Sheet Name:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.sheet_var = tk.StringVar(value=SHEET_NAME)
        ttk.Entry(file_frame, textvariable=self.sheet_var, width=20).grid(row=1, column=1, padx=5, pady=5, sticky="w")

        date_frame = ttk.LabelFrame(self, text="Fetch Transactions Since")
        date_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(date_frame, text="Enter Date:").grid(row=0, column=0, padx=5, pady=5, sticky="e")

        today = datetime.datetime.today()
        self.since_date_var = tk.StringVar(value=today.strftime("%d-%m-%Y"))
        self.since_date_entry = ttk.Entry(date_frame, textvariable=self.since_date_var, width=15)
        self.since_date_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        self.last_processed_label_var = tk.StringVar()
        last_processed = get_last_processed_time()
        if last_processed:
            self.last_processed_label_var.set(
                f"Last Updated: {last_processed.strftime('%d-%b-%Y %I:%M:%S %p')}"
            )
        else:
            self.last_processed_label_var.set("Last Run: None")

        self.last_run_label = ttk.Label(
            date_frame,
            textvariable=self.last_processed_label_var,
            style="LastRun.TLabel"
        )
        self.last_run_label.grid(row=1, column=0, columnspan=2, padx=5, pady=2, sticky="w")

        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=10, pady=5)

        self.fetch_btn = ttk.Button(btn_frame, text="Fetch UPI Transactions", command=self.on_fetch_clicked)
        self.fetch_btn.pack(side="left", padx=5)

        self.update_btn = ttk.Button(btn_frame, text="Update Excel", command=self.on_update_clicked)
        self.update_btn.pack(side="left", padx=5)

        self.quit_btn = ttk.Button(btn_frame, text="Quit", command=self.on_quit_clicked)
        self.quit_btn.pack(side="right", padx=5)

        self.progress_bar = ttk.Progressbar(btn_frame, orient="horizontal", mode="determinate")

        self.tree_frame = ttk.Frame(self)
        self.tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        columns = ("date", "party_name", "vpa_id", "amount", "category")
        self.tree = ttk.Treeview(self.tree_frame, columns=columns, show="headings", height=15)

        self.tree.heading("date", text="Date", anchor="center")
        self.tree.heading("party_name", text="Party Name", anchor="center")
        self.tree.heading("vpa_id", text="UPI ID", anchor="center")
        self.tree.heading("amount", text="Amount (Rs)", anchor="center")
        self.tree.heading("category", text="Category", anchor="center")

        self.tree.column("date", width=120, anchor="center")
        self.tree.column("party_name", width=150, anchor="center")
        self.tree.column("vpa_id", width=150, anchor="center")
        self.tree.column("amount", width=120, anchor="center")
        self.tree.column("category", width=120, anchor="center")

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        cat_frame = ttk.LabelFrame(self, text="")
        cat_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(cat_frame, text="Choose Category:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.single_combobox_var = tk.StringVar()
        self.single_combobox = ttk.Combobox(
            cat_frame,
            textvariable=self.single_combobox_var,
            values=list(EXPENSE_CATEGORIES.values()),
            state="readonly",
            width=25
        )
        self.single_combobox.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        self.apply_selected_btn = ttk.Button(cat_frame, text="Apply to Selected", command=self.on_apply_selected_clicked)
        self.apply_selected_btn.grid(row=0, column=2, padx=5, pady=5)

        self.apply_all_btn = ttk.Button(cat_frame, text="Apply to All", command=self.on_apply_all_clicked)
        self.apply_all_btn.grid(row=0, column=3, padx=5, pady=5)

        # Internals
        self.transactions = []
        self.category_choices = []
        self.fetch_thread = None
        self.total_emails = 0
        self.processed_emails = 0
        self.since_date_str = ""

    def on_quit_clicked(self):
        self.destroy()

    def on_fetch_clicked(self):
        if self.fetch_thread and self.fetch_thread.is_alive():
            messagebox.showwarning("Warning", "Fetching is already in progress.")
            return

        self.set_button_processing(self.fetch_btn)
        self.progress_bar.pack(side="left", padx=5)
        self.progress_bar["value"] = 0
        self.progress_bar.config(mode="determinate", maximum=1)

        # Clear old data
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.transactions.clear()
        self.category_choices.clear()

        self.fetch_thread = threading.Thread(target=self.fetch_transactions_in_thread)
        self.fetch_thread.start()

    def on_update_clicked(self):
        self.set_button_processing(self.update_btn)
        try:
            self.update_excel_gui()
        finally:
            self.set_button_normal(self.update_btn, "Update Excel")

    def on_apply_selected_clicked(self):
        self.set_button_processing(self.apply_selected_btn)
        try:
            selected_item_id = self.tree.focus()
            if not selected_item_id:
                messagebox.showwarning("No Selection", "Please select a row.")
                return

            row_index = int(selected_item_id)
            chosen_cat = self.single_combobox_var.get()
            if not chosen_cat:
                messagebox.showwarning("No Category", "Pick a category first.")
                return

            self.category_choices[row_index] = chosen_cat
            values = list(self.tree.item(row_index, "values"))
            values[-1] = chosen_cat
            self.tree.item(row_index, values=values)
        finally:
            self.set_button_normal(self.apply_selected_btn, "Apply to Selected")

    def on_apply_all_clicked(self):
        self.set_button_processing(self.apply_all_btn)
        try:
            chosen_cat = self.single_combobox_var.get()
            if not chosen_cat:
                messagebox.showwarning("No Category", "Pick a category first.")
                return

            for i in range(len(self.transactions)):
                self.category_choices[i] = chosen_cat
                values = list(self.tree.item(i, "values"))
                values[-1] = chosen_cat
                self.tree.item(i, values=values)
        finally:
            self.set_button_normal(self.apply_all_btn, "Apply to All")

    def set_button_processing(self, btn, label="Processing..."):
        btn.config(text=label, state="disabled")

    def set_button_normal(self, btn, original_text):
        btn.config(text=original_text, state="normal")

    # ---------- Fetch (Threaded) ----------
    def fetch_transactions_in_thread(self):
        mail = connect_gmail()
        if not mail:
            self.after(0, lambda: messagebox.showerror("Error", "Failed to connect to Gmail. Check credentials."))
            self.after(0, self.fetch_done)
            return

        # parse user date
        selected_date_str = self.since_date_var.get().strip()
        try:
            selected_date = datetime.datetime.strptime(selected_date_str, "%d-%m-%Y").date()
        except ValueError:
            selected_date = datetime.date.today()

        self.since_date_str = selected_date.strftime("%d-%b-%Y")

        search_query = f'(FROM "alerts@hdfcbank.net" SINCE "{self.since_date_str}")'
        try:
            result, data = mail.search(None, search_query)
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Error", f"Error searching mailbox:\n{e}"))
            self.after(0, lambda: self.fetch_done(mail))
            return

        if result != "OK":
            self.after(0, lambda: messagebox.showerror("Error", "Failed to search emails."))
            self.after(0, lambda: self.fetch_done(mail))
            return

        email_ids = data[0].split()
        self.total_emails = len(email_ids)
        self.after(0, lambda: self.progress_bar.config(maximum=self.total_emails, value=0))

        if not email_ids:
            self.after(0, lambda: self.fetch_done(mail, no_emails=True))
            return

        last_processed = get_last_processed_time()
        upi_pattern = r"Rs\.\s?(\d+\.\d{2})\s?has been debited .*? to VPA (\S+)\s(.+?) on (\d{2}-\d{2}-\d{2})"

        self.processed_emails = 0

        for e_id in email_ids:
            try:
                _, msg_data = mail.fetch(e_id, "(RFC822)")
                raw_email = msg_data[0][1]
                msg = email.message_from_bytes(raw_email)

                msg_date_hdr = msg["Date"]
                try:
                    msg_datetime = parsedate_to_datetime(msg_date_hdr)
                except:
                    msg_datetime = None

                # skip if <= last_processed
                if last_processed and msg_datetime and msg_datetime <= last_processed:
                    self.processed_emails += 1
                    self.after(0, self.increment_progress)
                    continue

                email_text = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        ct = part.get_content_type()
                        cd = str(part.get("Content-Disposition"))
                        if ct == "text/plain" and "attachment" not in cd:
                            email_text = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                            break
                        elif ct == "text/html" and not email_text:
                            email_text = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                else:
                    email_text = msg.get_payload(decode=True).decode("utf-8", errors="ignore")

                matches = re.findall(upi_pattern, email_text)
                parsed_transactions = []
                if matches:
                    for match in matches:
                        amount = float(match[0])
                        vpa_id = match[1]
                        party_name = match[2].strip().lower()
                        date_str = match[3]

                        txn = {
                            "date": date_str,
                            "amount": amount,
                            "vpa_id": vpa_id,
                            "party_name": party_name,
                            "email_datetime": msg_datetime
                        }
                        parsed_transactions.append(txn)

                if parsed_transactions:
                    self.after(0, lambda txns=parsed_transactions: self.add_transactions_to_ui(txns))

                self.processed_emails += 1
                self.after(0, self.increment_progress)

            except Exception as e2:
                print(f"Error processing email ID {e_id}: {e2}")

        self.after(0, lambda: self.fetch_done(mail))

    def add_transactions_to_ui(self, txns):
        for txn in txns:
            self.transactions.append(txn)
            self.category_choices.append("")
            row_index = len(self.transactions) - 1
            self.tree.insert(
                "",
                "end",
                iid=row_index,
                values=(
                    txn["date"],
                    txn["party_name"],
                    txn["vpa_id"],
                    txn["amount"],
                    ""
                )
            )

    def increment_progress(self):
        self.progress_bar["value"] = self.processed_emails

    def fetch_done(self, mail=None, no_emails=False):
        if mail:
            mail.logout()

        self.progress_bar.stop()
        self.progress_bar.pack_forget()
        self.set_button_normal(self.fetch_btn, "Fetch UPI Transactions")

        if no_emails:
            messagebox.showinfo("Info", f"No UPI transactions found since {self.since_date_str}.")
            return

        if self.total_emails > 0 and not self.transactions:
            last_processed = get_last_processed_time()
            if last_processed:
                messagebox.showinfo("Info", "No new transactions found since last run.")
            else:
                messagebox.showinfo("Info", "No valid UPI transactions found in the fetched emails.")
        else:
            messagebox.showinfo("Success", "Fetching completed successfully! Transactions are now listed below.")

    # ---------- Update Excel ----------
    def update_excel_gui(self):
        if not self.transactions:
            messagebox.showwarning("Warning", "No transactions to update.")
            return

        for i in range(len(self.transactions)):
            row_vals = list(self.tree.item(i, "values"))
            cat = row_vals[-1].strip() if row_vals[-1] else ""
            self.category_choices[i] = cat

        excel_file = self.excel_path_var.get().strip()
        sheet_name = self.sheet_var.get().strip()

        update_excel(self.transactions, self.category_choices, excel_file, sheet_name)

        for row in self.tree.get_children():
            self.tree.delete(row)
        self.transactions.clear()
        self.category_choices.clear()

        last_processed = get_last_processed_time()
        if last_processed:
            self.last_processed_label_var.set(
                f"Last Run: {last_processed.strftime('%d-%b-%Y %I:%M:%S %p')}"
            )

    def browse_excel(self):
        path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xlsb *.xls")],
        )
        if path:
            self.excel_path_var.set(path)

def main():
    app = ExpenseGUI()
    app.mainloop()

if __name__ == "__main__":
    main()
